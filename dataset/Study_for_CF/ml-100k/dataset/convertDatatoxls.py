#!/usr/bin/python
import openpyxl

wb=openpyxl.load_workbook('data.xlsx')
print wb.get_sheet_names()
sheet=wb.get_sheet_by_name('Sheet1')
print sheet['A1'].value

fl=open('u.data','r')
a=fl.readline()
print a
listrow=a.split()
print listrow

#user=int(listrow[0])
#item=int(listrow[1])
#rating=int(listrow[2])
#print user
#print item
#sheet.cell(row=user,column=item).value='yogi'


with open('u.data', 'r') as f:
  for line in f:
    read_data=line.split()
    user=int(read_data[0])
    movie=int(read_data[1])
    rating=int(read_data[2])
    sheet.cell(row=user,column=movie).value=rating
f.closed


#with open('u.data', 'r') as f:
#   print 'yogi'
#   read_data=f.read().split()
#   user=int(read_data[0])
#   movie=int(read_data[1])
 #  rating=int(read_data[2])
#   sheet.cell(row=user,column=movie).value=rating
#f.close


wb.save('data12.xlsx')
