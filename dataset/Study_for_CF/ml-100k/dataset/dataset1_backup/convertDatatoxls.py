#!/usr/bin/python
import openpyxl
import openpyxl.utils


wb=openpyxl.Workbook(data_only=True)
#sheet 0 for base data
sheet=wb.create_sheet(0)

sheet.title = "u1.base"
sheet.sheet_properties.tabColor = "00ffff"

with open('u1.base', 'r') as f:
  for line in f:
    read_data=line.split()
    user=int(read_data[0])
    movie=int(read_data[1])
    rating=int(read_data[2])
    sheet.cell(row=user,column=movie).value=rating
f.closed


#sheet 2 testing
sheet2=wb.create_sheet()

sheet2.title = "u1.test"
sheet2.sheet_properties.tabColor = "ffff00"

with open('u1.test', 'r') as f:
  for line in f:
    read_data=line.split()
    user=int(read_data[0])
    movie=int(read_data[1])
    rating=int(read_data[2])
    sheet2.cell(row=user,column=movie).value=rating
f.closed

# working with formulla
sh1=wb.get_sheet_by_name('Sheet')
sh1.cell(row=1,column=1).value='User'
sh1.cell(row=1,column=2).value='Mean of User'

#for i in range(2,944):
#   x="=AVERAGE(\'u1.base\'.A"+str((i-1))+":AMJ"+str((i-1))+")"
#   sh1.cell(row=i,column=2).set_explicit_value(value=x,data_type=openpyxl.cell.Cell.TYPE_FORMULA)



wb.save('data12.xlsx')
